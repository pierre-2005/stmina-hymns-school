from __future__ import annotations

import hashlib
import json
import os
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Optional
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from openpyxl import load_workbook

_cache: Dict[str, Any] = {}
_cache_key: tuple[str, int, int] | None = None


class ContentError(RuntimeError):
    """Raised when the content workbook cannot be loaded safely."""


def _parse_time_to_ms(value: Any) -> int:
    """Convert seconds, m:ss(.s), or h:mm:ss(.s) into milliseconds."""
    text = str(value or "").strip()
    if not text:
        return 0

    try:
        parts = text.split(":")
        if len(parts) == 1:
            return max(0, int(round(float(parts[0]) * 1000)))

        seconds = float(parts[-1])
        leading = [int(part) for part in parts[:-1]]
        if len(leading) == 1:
            total = leading[0] * 60 + seconds
        elif len(leading) == 2:
            total = leading[0] * 3600 + leading[1] * 60 + seconds
        else:
            raise ValueError("too many time components")
        return max(0, int(round(total * 1000)))
    except (TypeError, ValueError) as exc:
        raise ContentError(f"Invalid lyric timestamp '{text}'. Use formats such as 0:06.5 or 2:15.") from exc


def _truthy(value: Any, *, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _safe_float(value: Any, default: float = 1.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _read_table(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean(value) for value in rows[0]]
    output: list[dict[str, Any]] = []
    for excel_row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value is None or _clean(value) == "" for value in row):
            continue
        record = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        record["__row__"] = excel_row_number
        output.append(record)
    return output


def soundcloud_embed_url(raw_url: str, *, colour: str = "7A263A") -> str:
    """
    Accept either a normal public SoundCloud URL or an existing widget URL.
    The returned URL is ready for a responsive iframe.
    """
    url = _clean(raw_url)
    if not url:
        return ""

    parsed = urlparse(url)
    if parsed.netloc.lower() == "w.soundcloud.com" and parsed.path.startswith("/player"):
        params = dict(parse_qsl(parsed.query, keep_blank_values=True))
        params.setdefault("color", colour)
        params.setdefault("auto_play", "false")
        params.setdefault("hide_related", "true")
        params.setdefault("show_comments", "false")
        params.setdefault("show_user", "true")
        params.setdefault("show_reposts", "false")
        params.setdefault("show_teaser", "false")
        params.setdefault("visual", "false")
        return urlunparse(parsed._replace(query=urlencode(params)))

    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""

    query = urlencode(
        {
            "url": url,
            "color": colour,
            "auto_play": "false",
            "hide_related": "true",
            "show_comments": "false",
            "show_user": "true",
            "show_reposts": "false",
            "show_teaser": "false",
            "visual": "false",
        }
    )
    return f"https://w.soundcloud.com/player/?{query}"


def _normalise_site(site: Dict[str, Any]) -> Dict[str, Any]:
    languages = site.get("languages", [])
    default_fingerprint = "|".join(
        f"{language.get('code', '')}:{1 if language.get('default_on') else 0}" for language in languages
    )
    site["language_preferences_version"] = hashlib.sha1(default_fingerprint.encode("utf-8")).hexdigest()[:12]
    site.setdefault("content_warnings", [])
    return site


def _load_xlsx(content_path: str) -> Dict[str, Any]:
    workbook = load_workbook(content_path, data_only=True, read_only=True)
    warnings: list[str] = []

    site: Dict[str, Any] = {
        "site_title": "St. Mina Hymns School",
        "site_subtitle": "St. Mina Coptic Orthodox Church • Calgary, AB",
        "footer_text": "St. Mina Coptic Orthodox Church (Calgary)",
        "languages": [],
        "levels": [],
        "content_warnings": warnings,
    }

    for row in _read_table(workbook, "meta"):
        key = _clean(row.get("key"))
        if key:
            site[key] = row.get("value") if row.get("value") is not None else ""

    language_rows = _read_table(workbook, "languages")
    languages: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for original_index, row in enumerate(language_rows):
        code = _clean(row.get("code")).lower()
        if not code:
            warnings.append(f"languages row {row['__row__']} was ignored because code is blank.")
            continue
        if code in seen_codes:
            warnings.append(f"languages row {row['__row__']} duplicates code '{code}' and was ignored.")
            continue
        seen_codes.add(code)
        languages.append(
            {
                "code": code,
                "name": _clean(row.get("name")) or code,
                "is_rtl": _truthy(row.get("is_rtl"), default=False),
                "default_on": _truthy(row.get("default_on"), default=True),
                "sort": _safe_int(row.get("sort"), original_index * 10),
                "_row_order": original_index,
            }
        )
    languages.sort(key=lambda item: (item["sort"], item["_row_order"]))
    for language in languages:
        language.pop("_row_order", None)
    site["languages"] = languages

    levels_by_slug: dict[str, dict[str, Any]] = {}
    for row in _read_table(workbook, "levels"):
        if not _truthy(row.get("published"), default=True):
            continue
        slug = _clean(row.get("level_slug"))
        if not slug:
            warnings.append(f"levels row {row['__row__']} was ignored because level_slug is blank.")
            continue
        if slug in levels_by_slug:
            warnings.append(f"levels row {row['__row__']} duplicates level_slug '{slug}' and was ignored.")
            continue
        levels_by_slug[slug] = {
            "slug": slug,
            "name": _clean(row.get("level_name")) or slug,
            "description": _clean(row.get("level_description")),
            "sort": _safe_int(row.get("sort"), 0),
            "years": [],
        }

    years_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    years_by_slug: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for row in _read_table(workbook, "years"):
        if not _truthy(row.get("published"), default=True):
            continue
        level_slug = _clean(row.get("level_slug"))
        year_slug = _clean(row.get("year_slug"))
        if not level_slug or not year_slug:
            warnings.append(f"years row {row['__row__']} needs both level_slug and year_slug.")
            continue
        level = levels_by_slug.get(level_slug)
        if level is None:
            warnings.append(
                f"years row {row['__row__']} references unknown level_slug '{level_slug}' and was ignored."
            )
            continue
        key = (level_slug, year_slug)
        if key in years_by_key:
            warnings.append(
                f"years row {row['__row__']} duplicates year '{year_slug}' inside level '{level_slug}' and was ignored."
            )
            continue
        year = {
            "slug": year_slug,
            "level_slug": level_slug,
            "name": _clean(row.get("year_name")) or year_slug,
            "description": _clean(row.get("year_description")),
            "sort": _safe_int(row.get("sort"), 0),
            "hymns": [],
        }
        years_by_key[key] = year
        years_by_slug[year_slug].append(key)
        level["years"].append(year)

    hymns_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    hymns_by_slug: dict[str, list[tuple[str, str, str]]] = defaultdict(list)

    def resolve_year(row: dict[str, Any]) -> tuple[str, str] | None:
        year_slug = _clean(row.get("year_slug"))
        level_slug = _clean(row.get("level_slug"))
        if level_slug:
            key = (level_slug, year_slug)
            return key if key in years_by_key else None
        candidates = years_by_slug.get(year_slug, [])
        if len(candidates) == 1:
            return candidates[0]
        return None

    for row in _read_table(workbook, "hymns"):
        if not _truthy(row.get("published"), default=True):
            continue
        hymn_slug = _clean(row.get("hymn_slug"))
        year_key = resolve_year(row)
        if not hymn_slug or year_key is None:
            warnings.append(
                f"hymns row {row['__row__']} could not be linked. Use a unique year_slug, or add level_slug."
            )
            continue
        key = (*year_key, hymn_slug)
        if key in hymns_by_key:
            warnings.append(f"hymns row {row['__row__']} duplicates hymn_slug '{hymn_slug}' in that year.")
            continue
        hymn = {
            "slug": hymn_slug,
            "title": _clean(row.get("hymn_title")) or hymn_slug,
            "note": _clean(row.get("hymn_note")),
            "sort": _safe_int(row.get("sort"), 0),
            "recordings": [],
            "segments": [],
        }
        hymns_by_key[key] = hymn
        hymns_by_slug[hymn_slug].append(key)
        years_by_key[year_key]["hymns"].append(hymn)

    def resolve_hymn(row: dict[str, Any]) -> tuple[str, str, str] | None:
        hymn_slug = _clean(row.get("hymn_slug"))
        level_slug = _clean(row.get("level_slug"))
        year_slug = _clean(row.get("year_slug"))
        if level_slug and year_slug:
            key = (level_slug, year_slug, hymn_slug)
            return key if key in hymns_by_key else None
        candidates = hymns_by_slug.get(hymn_slug, [])
        if year_slug:
            candidates = [key for key in candidates if key[1] == year_slug]
        return candidates[0] if len(candidates) == 1 else None

    for row in _read_table(workbook, "recordings"):
        if not _truthy(row.get("published"), default=True):
            continue
        hymn_key = resolve_hymn(row)
        raw_url = _clean(row.get("soundcloud_url")) or _clean(row.get("url"))
        if hymn_key is None or not raw_url:
            warnings.append(
                f"recordings row {row['__row__']} was ignored. Check hymn_slug and the SoundCloud URL."
            )
            continue
        embed_url = soundcloud_embed_url(raw_url)
        if not embed_url:
            warnings.append(f"recordings row {row['__row__']} contains an invalid URL and was ignored.")
            continue
        hymns_by_key[hymn_key]["recordings"].append(
            {
                "label": _clean(row.get("label")) or "Recording",
                "url": raw_url,
                "embed_url": embed_url,
                "sort": _safe_int(row.get("sort"), 0),
            }
        )

    language_codes = [language["code"] for language in languages]
    for row in _read_table(workbook, "segments"):
        if not _truthy(row.get("published"), default=True):
            continue
        hymn_key = resolve_hymn(row)
        if hymn_key is None:
            warnings.append(f"segments row {row['__row__']} could not be linked to a hymn and was ignored.")
            continue
        try:
            start_ms = _parse_time_to_ms(row.get("t"))
        except ContentError as exc:
            warnings.append(f"segments row {row['__row__']}: {exc}")
            continue
        texts = {
            code: str(row.get(code))
            for code in language_codes
            if row.get(code) is not None and _clean(row.get(code))
        }
        if not texts:
            warnings.append(f"segments row {row['__row__']} has no lyric text and was ignored.")
            continue
        hymns_by_key[hymn_key]["segments"].append(
            {"t": _clean(row.get("t")) or "0:00", "start_ms": start_ms, "texts": texts}
        )

    for level in levels_by_slug.values():
        level["years"].sort(key=lambda year: (year["sort"], year["name"].casefold()))
        for year in level["years"]:
            year["hymns"].sort(key=lambda hymn: (hymn["sort"], hymn["title"].casefold()))
            for hymn in year["hymns"]:
                hymn["recordings"].sort(key=lambda recording: (recording["sort"], recording["label"].casefold()))
                hymn["segments"].sort(key=lambda segment: segment["start_ms"])

    site["levels"] = sorted(
        levels_by_slug.values(), key=lambda level: (level["sort"], level["name"].casefold())
    )
    return _normalise_site(site)


def _load_json(content_path: str) -> Dict[str, Any]:
    """Load the Content Manager JSON format and prepare only published content for the public site."""
    with open(content_path, "r", encoding="utf-8") as file:
        raw = json.load(file)

    site: Dict[str, Any] = {
        "site_title": _clean(raw.get("site_title")) or "St. Mina Hymns School",
        "site_subtitle": _clean(raw.get("site_subtitle")) or "St. Mina Coptic Orthodox Church • Calgary, AB",
        "footer_text": _clean(raw.get("footer_text")) or "St. Mina Coptic Orthodox Church (Calgary)",
        "languages": [],
        "levels": [],
        "content_warnings": [],
    }

    languages = []
    for index, language in enumerate(raw.get("languages", []) or []):
        if not isinstance(language, dict):
            continue
        code = _clean(language.get("code")).lower()
        if not code:
            continue
        languages.append(
            {
                "code": code,
                "name": _clean(language.get("name")) or code,
                "is_rtl": _truthy(language.get("is_rtl"), default=False),
                "default_on": _truthy(language.get("default_on"), default=True),
                "sort": _safe_int(language.get("sort"), index * 10),
            }
        )
    site["languages"] = sorted(languages, key=lambda item: (item["sort"], item["name"].casefold()))

    levels = []
    for level_index, level in enumerate(raw.get("levels", []) or []):
        if not isinstance(level, dict) or not _truthy(level.get("published"), default=True):
            continue
        out_level = {
            "slug": _clean(level.get("slug")),
            "name": _clean(level.get("name")) or _clean(level.get("slug")),
            "description": _clean(level.get("description")),
            "sort": _safe_int(level.get("sort"), level_index * 10),
            "years": [],
        }

        years = []
        for year_index, year in enumerate(level.get("years", []) or []):
            if not isinstance(year, dict) or not _truthy(year.get("published"), default=True):
                continue
            out_year = {
                "slug": _clean(year.get("slug")),
                "level_slug": out_level["slug"],
                "name": _clean(year.get("name")) or _clean(year.get("slug")),
                "description": _clean(year.get("description")),
                "sort": _safe_int(year.get("sort"), year_index * 10),
                "hymns": [],
            }

            hymns = []
            for hymn_index, hymn in enumerate(year.get("hymns", []) or []):
                if not isinstance(hymn, dict) or not _truthy(hymn.get("published"), default=True):
                    continue
                out_hymn = {
                    "slug": _clean(hymn.get("slug")),
                    "title": _clean(hymn.get("title")) or _clean(hymn.get("slug")),
                    "note": _clean(hymn.get("note")),
                    "sort": _safe_int(hymn.get("sort"), hymn_index * 10),
                    "recordings": [],
                    "segments": [],
                }

                for recording_index, recording in enumerate(hymn.get("recordings", []) or []):
                    if not isinstance(recording, dict) or not _truthy(recording.get("published"), default=True):
                        continue
                    raw_url = _clean(recording.get("url"))
                    embed_url = soundcloud_embed_url(raw_url)
                    if not raw_url or not embed_url:
                        continue
                    out_hymn["recordings"].append(
                        {
                            "label": _clean(recording.get("label")) or "Recording",
                            "url": raw_url,
                            "embed_url": embed_url,
                            "sort": _safe_int(recording.get("sort"), recording_index * 10),
                        }
                    )

                for segment_index, segment in enumerate(hymn.get("segments", []) or []):
                    if not isinstance(segment, dict) or not _truthy(segment.get("published"), default=True):
                        continue
                    texts = {
                        _clean(code).lower(): str(value)
                        for code, value in (segment.get("texts") or {}).items()
                        if _clean(code) and value is not None and _clean(value)
                    }
                    if not texts:
                        continue
                    timestamp = _clean(segment.get("t")) or "0:00"
                    out_hymn["segments"].append(
                        {
                            "t": timestamp,
                            "start_ms": _parse_time_to_ms(timestamp),
                            "texts": texts,
                            "sort": _safe_int(segment.get("sort"), segment_index * 10),
                        }
                    )

                out_hymn["recordings"].sort(key=lambda item: (item["sort"], item["label"].casefold()))
                out_hymn["segments"].sort(key=lambda item: (item["start_ms"], item["sort"]))
                hymns.append(out_hymn)

            out_year["hymns"] = sorted(hymns, key=lambda item: (item["sort"], item["title"].casefold()))
            years.append(out_year)

        out_level["years"] = sorted(years, key=lambda item: (item["sort"], item["name"].casefold()))
        levels.append(out_level)

    site["levels"] = sorted(levels, key=lambda item: (item["sort"], item["name"].casefold()))
    return _normalise_site(site)


def load_site(content_path: str) -> Dict[str, Any]:
    global _cache, _cache_key

    path = Path(content_path)
    if not path.exists():
        raise ContentError(f"Content file not found: {content_path}")

    stat = path.stat()
    key = (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
    if _cache and _cache_key == key:
        return _cache

    site = _load_xlsx(content_path) if path.suffix.lower() == ".xlsx" else _load_json(content_path)
    _cache = site
    _cache_key = key
    return site


def find_level(site: Dict[str, Any], level_slug: str) -> Optional[Dict[str, Any]]:
    return next((level for level in site.get("levels", []) if level.get("slug") == level_slug), None)


def find_year(level: Dict[str, Any], year_slug: str) -> Optional[Dict[str, Any]]:
    return next((year for year in level.get("years", []) if year.get("slug") == year_slug), None)


def find_hymn(year: Dict[str, Any], hymn_slug: str) -> Optional[Dict[str, Any]]:
    return next((hymn for hymn in year.get("hymns", []) if hymn.get("slug") == hymn_slug), None)


def flatten_hymns(site: Dict[str, Any]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for level in site.get("levels", []):
        for year in level.get("years", []):
            for hymn in year.get("hymns", []):
                output.append(
                    {
                        "slug": hymn["slug"],
                        "title": hymn["title"],
                        "level_slug": level["slug"],
                        "level_name": level["name"],
                        "year_slug": year["slug"],
                        "year_name": year["name"],
                    }
                )
    return output
